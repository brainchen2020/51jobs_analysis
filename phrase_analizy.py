# -*- coding: utf-8 -*-
"""
Created on Fri Sep  7 19:00:08 2018

@author: brainbot
"""

import re
from openpyxl import load_workbook
import pandas as pd
from textrank4zh import TextRank4Keyword
from matplotlib.font_manager import FontProperties
import seaborn as sns
from datetime import datetime

def translate(str):
    # print("origon data:"+str.decode("utf-8","ignore"))
    line = str.strip().decode('utf-8','ignore') #解码成unicode
    p2 = re.compile(u'[^\u4e00-\u9fa5]')  # 中文的编码范围是：\u4e00到\u9fa5
    outStr= "".join(p2.split(line)).strip()
    # print("translate done")
    return outStr
def get_xlsx_columns(wb,index=0):
    sheet = wb.active
    value_reg=""
    for item in range(sheet.max_row):
        if(index == 6):
            values=list(sheet.columns)[index][item].value.split("职能类别：")
            values[0]= ("".join(values[0])).encode('utf-8')
            
        else:
            values=list(sheet.columns)[index][item].value
            values= "".join(values)
        value_reg=value_reg+translate(values[0])
    print("get_xlsx_columns done value_reg",value_reg)
    return value_reg
def analyze_word(text):
    arr=""
    tr4w = TextRank4Keyword()
    tr4w.analyze(text=text, lower=True, window=2)
    for words in tr4w.words_all_filters:
        arr=arr+'/'.join(words)
    print("analyze_word done arr",arr)
    return arr

def plot_sns(text):
    ##list(set())去字符串重复
    li=list(set(text.split("/")))
    counte_phrase=[]
    #counte_phrase.append(["pharse","times"])
    counte_phrase_phrase=[]
    counte_phrase_times=[]
    for i in range(1000):
        ##验证过滤掉单个词。
        if(len(li[i]) != 1):
            counte_phrase.append([li[i],text.count(li[i])])
            counte_phrase_times.append(text.count(li[i]))
            counte_phrase_phrase.append(li[i])
#    counte_phrase.append([li[i],text.count(li[i])])
#    pdf=pd.DataFrame().append(counte_phrase)
#print(pdf.sort_values(1))
#print(sorted(counte_phrase,key=lambda counte_phrase:counte_phrase[1],reverse=True))
    pdf_=pd.DataFrame({"phrase":counte_phrase_phrase,
                       "times":counte_phrase_times}).sort_values("times",ascending=False)
    pdf_.to_csv('phrase_reg.csv',header=True,index=False)
    zhfont=FontProperties(fname=r'C:\Windows\Fonts\simhei.ttf',size=14)
    sns.set(font=zhfont.get_name())
    sns.barplot("phrase", "times", palette="RdBu_r", data=pdf_.head(10))
    sns.barplot("phrase", "times", palette="RdBu_r", data=pdf_.head(-10))

t1=datetime.now()
print(t1)
wb = load_workbook("51jobs.xlsx")
text=analyze_word(get_xlsx_columns(wb,6))
plot_sns(text)

t2=datetime.now()-t1
print(t2)
