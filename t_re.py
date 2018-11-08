# -*- coding: utf-8 -*-
import re
from openpyxl import load_workbook
from collections import Counter
import matplotlib.pyplot as plt
import numpy as np
from textrank4zh import TextRank4Keyword, TextRank4Sentence
tr4w=TextRank4Keyword()

wb = load_workbook("51jobs.xlsx")
sheet_1 = wb.sheetnames
sheet = wb.active

##格式化输出。。只显示中文
def translate(str):
    # print("origon data:"+str.decode("utf-8","ignore"))
    line = str.strip().decode('utf-8','ignore') #解码成unicode
    p2 = re.compile(u'[^\u4e00-\u9fa5]')  # 中文的编码范围是：\u4e00到\u9fa5
    outStr= "".join(p2.split(line)).strip()
    return outStr

#按行输出行
# for row in sheet.rows:
#     for cell in row:
#         print(cell.value)
str_arr=""
##按行输出列
for item in range(sheet.max_row):
    descriptions=list(sheet.columns)[6][item].value.split("职能类别：")
    descriptions[0]= "".join(descriptions[0])
    print(descriptions[0])
    # for s in tr4w.sentences:
    #     print(s)
    # print()
    # for words in tr4w.words_no_filter:
    #     print("/".join(words))
    # print()
    for words in tr4w.words_no_stop_words:
        print("/".join(words))
    # str_arr=str_arr+translate(descriptions[0].encode('utf-8'))

# with open("wc_clear.txt",'w') as f :
#     f.write(str_arr)
#     print("wrote done")
# print(Counter(str_arr))
# print(max(list(Counter(str_arr).values())))
# print(list(zip(Counter(str_arr).keys())))

##画图
# max_value = max(list(Counter(str_arr).values()))
# min_value = min(list(Counter(str_arr).values()))
# plt.hist(Counter(str_arr).keys(),
#          bins=np.arange(min_value,max_value))
# plt.title("最大值")
# plt.xlabel("单字")
# plt.ylabel("出现次数")
# plt.show()



